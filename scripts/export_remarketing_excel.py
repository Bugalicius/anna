#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
scripts/export_remarketing_excel.py

Exporta o relatorio_remarketing.json para Excel (nome + numero + categoria).

Uso:
  pip install openpyxl
  python scripts/export_remarketing_excel.py
  python scripts/export_remarketing_excel.py --output remarketing.xlsx
"""

from __future__ import annotations

import json
import argparse
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    raise SystemExit("Instale openpyxl: pip install openpyxl")

SCRIPTS = Path(__file__).resolve().parent
JSON_FILE = SCRIPTS / "relatorio_remarketing.json"

CATEGORIAS = [
    ("ex_paciente",  "Ex-pacientes",   "FFF2CC"),
    ("quase_marcou", "Quase marcaram", "D9EAD3"),
    ("lead_frio",    "CFE2F3",         "CFE2F3"),
]

CATEGORIAS = [
    ("ex_paciente",  "Ex-pacientes",   "FFF2CC"),
    ("quase_marcou", "Quase marcaram", "D9EAD3"),
    ("lead_frio",    "Leads frios",    "CFE2F3"),
]

LABEL = {
    "ex_paciente":  "Ex-paciente",
    "quase_marcou": "Quase marcou",
    "lead_frio":    "Lead frio",
}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", default="scripts/remarketing.xlsx")
    args = parser.parse_args()

    data = json.loads(JSON_FILE.read_text(encoding="utf-8"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Remarketing"

    # Cabecalho
    headers = ["Nome", "Numero", "Categoria"]
    header_fill = PatternFill("solid", fgColor="4A86C8")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for key, label, color in CATEGORIAS:
        fill = PatternFill("solid", fgColor=color)
        for entry in data.get(key, []):
            numero = entry["numero"]
            if numero == "+0":
                continue
            ws.cell(row=row, column=1, value=entry["nome"]).fill = fill
            ws.cell(row=row, column=2, value=numero).fill = fill
            ws.cell(row=row, column=3, value=LABEL[key]).fill = fill
            row += 1

    # Largura das colunas
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18

    out = Path(args.output)
    wb.save(out)
    print(f"Excel salvo em: {out}")
    print(f"  Total de linhas: {row - 2}")


if __name__ == "__main__":
    main()
